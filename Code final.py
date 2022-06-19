import numpy
import numpy as np
import random
import openpyxl as xl

def angle(V,R):
    D1=np.sqrt((V[0]**2)+(V[1]**2))
    D2 = np.sqrt((R[0] ** 2) + (R[1] ** 2))
    alpha=np.arccos(((V[0]*R[0])+(V[1]*R[1]))/(D1*D2))
    return alpha
def zone(VOB):
    if(VOB[0]>0 and VOB[1]>0 ):
        ob='B'
    else:
        if (VOB[0] > 0 and VOB[1] < 0):
            ob='D'
        else:
            if (VOB[0] < 0 and VOB[1] > 0):
                ob='A'
            else:
                if (VOB[0] < 0 and VOB[1] < 0):
                    ob='C'
                else:
                    if (VOB[0] == 0 and VOB[1] > 0):
                        ob='AB'
                    else:
                        if (VOB[0] > 0 and VOB[1]==0):
                            ob='BD'
                        else:
                            if (VOB[0]==0 and VOB[1] < 0):
                                ob='CD'
                            else:
                                if (VOB[0] < 0 and VOB[1] ==0):
                                    ob='AC'
                                else:
                                    ob='0'
    return ob
def zone_travaille(VOB,l):
    if(np.sqrt((VOB[0]**2)+(VOB[1]**2))>(2*l)):
        return 1
    else:
        return 0
def Resolution_inverse(Px, Py, l):
    Tet1 = [0, 0]
    Tet2 = [0, 0]
    Tet2[0] = numpy.arccos((((Px ** 2) + (Py ** 2)) / (2 * (l ** 2))) - 1)
    Tet2[1] = -Tet2[0]

    for i in range(2):
        Tet1[i] = (numpy.arctan(Px / Py)) - (Tet2[i] / 2)
    return Tet1, Tet2
def filtre_zone(Px,Py,VOB,Te1):
    Tet1=[0,0,0,0,0,0]
    Tet1[0] =Te1[0]
    Tet1[1] =Te1[1]
    Tet1[2] = Tet1[0] - 2 * np.pi
    Tet1[3] = Tet1[0] + 2 * np.pi
    Tet1[4] = Tet1[1] - 2 * np.pi
    Tet1[5] = Tet1[1] + 2 * np.pi
    VP=np.zeros(2)
    VP[0]=Px
    VP[1] =Py
    while (True):
        test = len(Tet1)
        for i in range(len(Tet1)):
            if (abs(Tet1[i]) >= 2 * np.pi):
                del Tet1[i]
                break
        if (test == len(Tet1)):
            break
    if(zone(VOB)!= zone(VP)):
        if(zone(VOB)=='B' or zone(VOB)=='AB'):
            while (True):
                test = len(Tet1)
                for i in range(len(Tet1)):
                    if (Tet1[i] >= 0):
                        del Tet1[i]

                        break
                if (test == len(Tet1)):
                    break
        if (zone(VOB) == 'D' or zone(VOB)=='CD'):
            while (True):
                test = len(Tet1)
                for i in range(len(Tet1)):
                    if (Tet1[i] <= 0):
                        del Tet1[i]
                        break
                if (test == len(Tet1)):
                    break

        if(zone(VOB) == 'C' and zone(VP)== 'D'):
            while (True):
                test = len(Tet1)
                for i in range(len(Tet1)):
                    if (Tet1[i] > 0 or abs(Tet1[i]) > (np.pi / 2)):
                        del Tet1[i]
                        break
                if (test == len(Tet1)):
                    break
        if (zone(VOB) == 'A' and zone(VP) == 'B'):
            while (True):
                test = len(Tet1)
                for i in range(len(Tet1)):
                    if (Tet1[i] < 0 or abs(Tet1[i]) > (np.pi / 2)):
                        del Tet1[i]
                        break
                if (test == len(Tet1)):
                    break
    return Tet1
def Project1(VOB,Tet1):
    A=(VOB[0]**2)+(VOB[1]**2)
    C=((VOB[0]*np.cos(Tet1))+(VOB[1]*np.sin(Tet1)))**2
    B=np.sqrt(A-C)
    return B
def Project2(VOB,Tet1,Tet2,l):
    A=(VOB[0]**2)+(VOB[1]**2)
    C=(((np.cos(Tet1+Tet2)+np.cos(Tet1))*VOB[0])+((np.sin(Tet1+Tet2)+np.sin(Tet1))*VOB[1]))/np.sqrt(2*(1+np.cos(Tet2)))
    B=np.sqrt(A-(C**2))
    return B
def projection_mouvement_seri(Tet2,Tet1,VOB,l,r):
    # vector O1O2
    V02 = np.zeros(2)
    # vector OzO1
    V01 = np.zeros(2)
    Ts=[1,1,1,1,1,1,1,1,1,1,1,1]
    Te1=0
    Te2=0
    beta=Tet1
    #Paralléle
    # Tet11 and Tet21
    while(True):
        V01[0] = l * (np.cos(Te1))
        V01[1] = l * (np.sin(Te1))
        V02[0] = l * np.cos(Te1+ Te2) + l * (np.cos(Te1))
        V02[1] = l * np.sin(Te1+Te2) + l * (np.sin(Te1))
        alpha = angle(VOB, V01)
        alpha1 = angle(VOB, V01)
        B0 = Project1(VOB, Te1)
        B1 = Project2(VOB,Te1, Te2, l)

        if ((B1 < r and abs(alpha1) < 1.58) or (B0 < r and abs(alpha) < 1.58)):
            Ts[0] = 0
            break

        if(abs(Te1) >= abs(Tet1[0]) and abs(Te2) >= abs(Tet2[0]) ):
            break
        if(Tet1[0]>=0):
            if(Te1 <= Tet1[0]):
                Te1=Te1+0.001

        else:
            if(Te1 >= Tet1[0]):
                Te1=Te1-0.001

        if(Tet2[0]>=0):
            if(Te2 <= Tet2[0]):
                Te2=Te2+0.001

        else:
            if(Te2 >= Tet2[0]):
                Te2=Te2-0.001

    Te1 = 0
    Te2 = 0
    # Tet11 and Tet22
    while (True):
        V01[0] = l * (np.cos(Te1))
        V01[1] = l * (np.sin(Te1))
        V02[0] = l * np.cos(Te1 + Te2) + l * (np.cos(Te1))
        V02[1] = l * np.sin(Te1 + Te2) + l * (np.sin(Te1))
        alpha = angle(VOB, V01)
        alpha1 = angle(VOB, V01)
        B0 = Project1(VOB, Te1)
        B1 = Project2(VOB, Te1, Te2, l)

        if ((B1 < r and abs(alpha1) < 1.58) or (B0 < r and abs(alpha) < 1.58)):
            Ts[1] = 0
            break
        if (abs(Te1) >= abs(Tet1[0]) and abs(Te2) >= abs(Tet2[1])):
            break
        if (Tet1[0] >= 0):
            if (Te1 <= Tet1[0]):
                Te1 = Te1 + 0.001

        else:
            if (Te1 >= Tet1[0]):
                Te1 = Te1 - 0.001


        if (Tet2[1] >= 0):
            if (Te2 <= Tet2[1]):
                Te2 = Te2 + 0.001

        else:
            if (Te2 >= Tet2[1]):
                Te2 = Te2 - 0.001

    Te1 = 0
    Te2 = 0
    if(len(Tet1)>=2):
        # Tet12 and Tet21
        while (True):
            V01[0] = l * (np.cos(Te1))
            V01[1] = l * (np.sin(Te1))
            V02[0] = l * np.cos(Te1 + Te2) + l * (np.cos(Te1))
            V02[1] = l * np.sin(Te1 + Te2) + l * (np.sin(Te1))
            alpha = angle(VOB, V01)
            alpha1 = angle(VOB, V01)
            B0 = Project1(VOB, Te1)
            B1 = Project2(VOB, Te1, Te2, l)
            if ((B1 < r and abs(alpha1) < 1.58) or (B0 < r and abs(alpha) < 1.58)):
                Ts[2] = 0
                break
            if (abs(Te1) >= abs(Tet1[1]) and abs(Te2) >= abs(Tet2[0])):
                break
            if (Tet1[1] >= 0):
                if (Te1 <= Tet1[1]):
                    Te1 = Te1 + 0.001
            else:
                if (Te1 >= Tet1[1]):
                    Te1 = Te1 - 0.001

            if (Tet2[0] >= 0):
                if (Te2 <= Tet2[0]):
                    Te2 = Te2 + 0.001
            else:
                if (Te2 >= Tet2[0]):
                    Te2 = Te2 - 0.001
        Te1 = 0
        Te2 = 0
        # Tet12 and Tet22
        while (True):
            V01[0] = l * (np.cos(Te1))
            V01[1] = l * (np.sin(Te1))
            V02[0] = l * np.cos(Te1 + Te2) + l * (np.cos(Te1))
            V02[1] = l * np.sin(Te1 + Te2) + l * (np.sin(Te1))
            alpha = angle(VOB, V01)
            alpha1 = angle(VOB, V01)
            B0 = Project1(VOB, Te1)
            B1 = Project2(VOB, Te1, Te2, l)
            if ((B1 < r and abs(alpha1) < 1.58) or (B0 < r and abs(alpha) < 1.58)):
                Ts[3] = 0
                break
            if (abs(Te1) >= abs(Tet1[1]) and abs(Te2) >= abs(Tet2[1])):
                break
            if (Tet1[1] >= 0):
                if (Te1 <= Tet1[1]):
                    Te1 = Te1 + 0.001
            else:
                if (Te1 >= Tet1[1]):
                    Te1 = Te1 - 0.001

            if (Tet2[1] >= 0):
                if (Te2 <= Tet2[1]):
                    Te2 = Te2 + 0.001
            else:
                if (Te2 >= Tet2[1]):
                    Te2 = Te2 - 0.001


    if(len(Tet1)>=3):
        Te1 = 0
        Te2 = 0
        # Tet13 and Tet21
        while (True):
            V01[0] = l * (np.cos(Te1))
            V01[1] = l * (np.sin(Te1))
            V02[0] = l * np.cos(Te1 + Te2) + l * (np.cos(Te1))
            V02[1] = l * np.sin(Te1 + Te2) + l * (np.sin(Te1))
            alpha = angle(VOB, V01)
            alpha1 = angle(VOB, V01)
            B0 = Project1(VOB, Te1)
            B1 = Project2(VOB, Te1, Te2, l)
            if ((B1 < r and abs(alpha1) < 1.58) or (B0 < r and abs(alpha) < 1.58)):
                Ts[4] = 0
                break
            if (abs(Te1) >= abs(Tet1[2]) and abs(Te2) >= abs(Tet2[0])):
                break
            if (Tet1[2] >= 0):
                if (Te1 <= Tet1[2]):
                    Te1 = Te1 + 0.001
            else:
                if (Te1 >= Tet1[2]):
                    Te1 = Te1 - 0.001

            if (Tet2[0] >= 0):
                if (Te2 <= Tet2[0]):
                    Te2 = Te2 + 0.001
            else:
                if (Te2 >= Tet2[0]):
                    Te2 = Te2 - 0.001

        Te1 = 0
        Te2 = 0
        # Tet13 and Tet22
        while (True):
            V01[0] = l * (np.cos(Te1))
            V01[1] = l * (np.sin(Te1))
            V02[0] = l * np.cos(Te1 + Te2) + l * (np.cos(Te1))
            V02[1] = l * np.sin(Te1 + Te2) + l * (np.sin(Te1))
            alpha = angle(VOB, V01)
            alpha1 = angle(VOB, V01)
            B0 = Project1(VOB, Te1)
            B1 = Project2(VOB, Te1, Te2, l)
            if ((B1 < r and abs(alpha1) < 1.58) or (B0 < r and abs(alpha) < 1.58)):
                Ts[5] = 0
                break
            if (abs(Te1) >= abs(Tet1[2]) and abs(Te2) >= abs(Tet2[1])):
                break
            if (Tet1[2] >= 0):
                if (Te1 <= Tet1[2]):
                    Te1 = Te1 + 0.001
            else:
                if (Te1 >= Tet1[2]):
                    Te1 = Te1 - 0.001

            if (Tet2[1] >= 0):
                if (Te2 <= Tet2[1]):
                    Te2 = Te2 + 0.001
            else:
                if (Te2 >= Tet2[1]):
                    Te2 = Te2 - 0.001

    if(len(Tet1)>=4):
        Te1 = 0
        Te2 = 0
        # Serie Tet14 and Tet21
        while (True):
            V01[0] = l * (np.cos(Te1))
            V01[1] = l * (np.sin(Te1))
            V02[0] = l * np.cos(Te1 + Te2) + l * (np.cos(Te1))
            V02[1] = l * np.sin(Te1 + Te2) + l * (np.sin(Te1))
            alpha = angle(VOB, V01)
            alpha1 = angle(VOB, V01)
            B0 = Project1(VOB, Te1)
            B1 = Project2(VOB, Te1, Te2, l)
            if ((B1 < r and abs(alpha1) < 1.58) or (B0 < r and abs(alpha) < 1.58)):
                Ts[6] = 0
                break
            if (abs(Te1) >= abs(Tet1[3]) and abs(Te2) >= abs(Tet2[0])):
                break
            if (Tet1[3] >= 0):
                if (Te1 <= Tet1[3]):
                    Te1 = Te1 + 0.001
            else:
                if (Te1 >= Tet1[3]):
                    Te1 = Te1 - 0.001

            if (Tet2[0] >= 0):
                if (Te2 <= Tet2[0]):
                    Te2 = Te2 + 0.001
            else:
                if (Te2 >= Tet2[0]):
                    Te2 = Te2 - 0.001

        Te1 = 0
        Te2 = 0
        # Tet14 and Tet22
        while (True):
            V01[0] = l * (np.cos(Te1))
            V01[1] = l * (np.sin(Te1))
            V02[0] = l * np.cos(Te1 + Te2) + l * (np.cos(Te1))
            V02[1] = l * np.sin(Te1 + Te2) + l * (np.sin(Te1))
            alpha = angle(VOB, V01)
            alpha1 = angle(VOB, V01)
            B0 = Project1(VOB, Te1)
            B1 = Project2(VOB, Te1, Te2, l)
            if ((B1 < r and abs(alpha1) < 1.58) or (B0 < r and abs(alpha) < 1.58)):
                Ts[7] = 0
                break
            if (abs(Te1) >= abs(Tet1[3]) and abs(Te2) >= abs(Tet2[1])):
                break
            if (Tet1[3] >= 0):
                if (Te1 <= Tet1[3]):
                    Te1 = Te1 + 0.001
            else:
                if (Te1 >= Tet1[3]):
                    Te1 = Te1 - 0.001

            if (Tet2[1] >= 0):
                if (Te2 <= Tet2[1]):
                    Te2 = Te2 + 0.001
            else:
                if (Te2 >= Tet2[1]):
                    Te2 = Te2 - 0.001
    c1=0
    c2 = 0
    c3 = 0
    c4 = 0
    if(Ts[0]==0 and Ts[1]==0):
      c1=c1+1
      del Tet1[0]

    if(Ts[2]==0 and Ts[3]==0):
        c2=c2+1
        del Tet1[1-c1]

    if(Ts[4]==0 and Ts[5]==0):
        c3=c3+1
        del Tet1[2-c1-c2]
    if(Ts[6]==0 and Ts[7]==0):
        c4=c4+1
        del Tet1[3-c1-c2-c3]
    if((Ts[1]==0 and Ts[3]==0 and Ts[5]==0 and Ts[7]==0)):
      del Tet2[1]
    if((Ts[0]==0 and Ts[2]==0 and Ts[4]==0 and Ts[6]==0)):
      del Tet2[0]
    to=np.zeros(4)
    o = 0
    if (not Tet1 and not Tet2):
        o = 1
    if (o == 0):
        for i in range(len(Tet1)):
            x1 = l * np.cos(Tet2[0] + Tet1[i]) + l * (np.cos(Tet1[i]))
            y1 = l * np.sin(Tet2[0] + Tet1[i]) + l * (np.sin(Tet1[i]))
            if (x1!=Px and y1!=Py):
                to[i]=1
    o = 0
    if (not Tet1):
        o = 1
    if(o==0 and (len(Tet2)==2)):
        for i in range(len(Tet1)):
            print(Tet1,Tet2)
            x1 = l * np.cos(Tet2[1] + Tet1[i]) + l * (np.cos(Tet1[i]))
            y1 = l * np.sin(Tet2[1] + Tet1[i]) + l * (np.sin(Tet1[i]))
            if (x1!=Px and y1!=Py):
                to[i+2]=1
        if(to[0]!=to[2] and to[1]!=to[3]):
            ab=Tet2[0]
            Tet2[0]=Tet2[1]
            Tet2[0]=ab
    return Tet1,Tet2

Px=100
Py=100
VO=np.zeros(2)
VO[0]=50
VO[1]=50
r=10
l=100
a=0
wb = xl.load_workbook('test.xlsx')
sheet = wb['Sheet1']
k=0for i in range(1000):
    print('iteration',i)
    Px=random.randint(0,20)
    Py = random.randint(1,20)
    if(i%2==0):
        Py = random.randint(1, 20)
    while (True):
        Pox = random.randint(1, 20)
        Poy = random.randint(1, 20)
        if (Pox != Px or Poy != Py):
            break
    VOB = np.zeros(2)
    VOB[0] = Pox
    VOB[1] = Poy
    a = zone_travaille(VO, l)
    [Tet1, Tet2] = Resolution_inverse(Px, Py, l)
    if (a == 1):
       del Tet1[1]
       del Tet2[1]
    else:
        Tet1 = filtre_zone(Px, Py, VO, Tet1)
        [Tet1, Tet2] = projection_mouvement_seri(Tet2, Tet1, VOB, l, r)

    o=0
    if(not Tet1):
        o=1
    if(o==0 ):
        ################
        Px_t = sheet.cell(1, k + 2)
        Px_t.value =(l * np.cos(Tet2[0]+ Tet1[0])+l * (np.cos(Tet1[0])))
        Py_t = sheet.cell(2, k + 2)
        Py_t.value = (l * np.sin(Tet2[0] + Tet1[0])+l * (np.sin(Tet1[0])))
        #################
        #################
        Pox_t = sheet.cell(3, k + 2)
        Pox_t.value = Pox
        Poy_t = sheet.cell(4, k + 2)
        Poy_t.value = Poy
        #################
        print(Tet1, Tet2)
        for j in range(len(Tet1)):
            Tet11 = sheet.cell(5 + j, k + 2)
            Tet11.value = Tet1[j]

        for j in range(len(Tet2)):
            Tet22 = sheet.cell(8 + j, k + 2)
            Tet22.value = Tet2[j]
        k=k+1
wb.save('base22.xlsx')



